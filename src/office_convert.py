"""Office document conversion utilities.

Provides converters for:
- DOCX → Markdown (via Docling Slim)
- PPTX → Markdown (via Docling Slim)
- XLSX → Markdown (via openpyxl)

Each converter returns the Markdown string and metadata about the conversion.
"""

from __future__ import annotations

import hashlib
import logging
import os
import tempfile
import time
import signal
import shutil
import threading
import multiprocessing
import queue
import sys
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# Timeout for document conversion (seconds)
from .config import OFFICE_MIN_FREE_DISK_MB, OFFICE_PARSE_TIMEOUT_SECONDS

class OfficeConversionError(RuntimeError):
    """Recoverable Office conversion failure with a stable reason."""


class PptxPreviewFileTooLargeError(RuntimeError):
    """The conversion service rejected a PPTX because it exceeded its limit."""

class _ParseTimeout(Exception):
    pass

def _ensure_disk_space(path: Path) -> None:
    usage = shutil.disk_usage(path.parent)
    required = OFFICE_MIN_FREE_DISK_MB * 1024 * 1024
    if usage.free < required:
        raise OfficeConversionError(f"office_disk_space_low: free={usage.free} required={required}")

def _run_with_timeout(func: Any, *args: Any, **kwargs: Any) -> Any:
    timeout = OFFICE_PARSE_TIMEOUT_SECONDS
    if timeout <= 0:
        raise OfficeConversionError("office_parse_timeout: timeout must be positive")
    can_alarm = hasattr(signal, "SIGALRM") and threading.current_thread() is threading.main_thread()
    previous = None
    def _alarm(_signum: int, _frame: Any) -> None:
        raise _ParseTimeout
    started = time.monotonic()
    try:
        if can_alarm:
            previous = signal.signal(signal.SIGALRM, _alarm)
            signal.setitimer(signal.ITIMER_REAL, timeout)
        result = func(*args, **kwargs)
        if time.monotonic() - started > timeout:
            raise _ParseTimeout
        return result
    except _ParseTimeout as exc:
        raise OfficeConversionError(f"office_parse_timeout: {timeout}s") from exc
    finally:
        if can_alarm:
            signal.setitimer(signal.ITIMER_REAL, 0)
            signal.signal(signal.SIGALRM, previous)

def _docling_process_entry(path: str, output: Any) -> None:
    try:
        from docling.document_converter import DocumentConverter
        result = DocumentConverter().convert(path)
        output.put(("ok", result.document.export_to_markdown()))
    except BaseException as exc:
        output.put(("error", f"{type(exc).__name__}: {exc}"))

def _xlsx_process_entry(path: str, output: Any) -> None:
    try:
        output.put(("ok", _convert_xlsx_to_markdown_impl(Path(path))))
    except BaseException as exc:
        output.put(("error", f"{type(exc).__name__}: {exc}"))

def _run_conversion_process(target: Any, path: Path) -> Any:
    timeout = OFFICE_PARSE_TIMEOUT_SECONDS
    if timeout <= 0:
        raise OfficeConversionError("office_parse_timeout: timeout must be positive")
    context = multiprocessing.get_context("spawn")
    output = context.Queue(maxsize=1)
    process = context.Process(target=target, args=(str(path), output), daemon=True)
    deadline = time.monotonic() + timeout
    process.start()
    try:
        while True:
            remaining = deadline - time.monotonic()
            if remaining <= 0:
                raise OfficeConversionError(f"office_parse_timeout: {timeout}s")
            try:
                status, payload = output.get(timeout=min(remaining, 0.1))
                break
            except queue.Empty as exc:
                if not process.is_alive():
                    process.join()
                    raise OfficeConversionError(
                        f"office_parse_failed: exit={process.exitcode}"
                    ) from exc

        process.join(max(0, deadline - time.monotonic()))
        if process.is_alive():
            raise OfficeConversionError(f"office_parse_timeout: {timeout}s")
    finally:
        if process.is_alive():
            process.terminate()
            process.join(5)
            if process.is_alive():
                process.kill()
                process.join()
        output.close()
        output.join_thread()
    if status != "ok":
        raise OfficeConversionError(f"office_parse_failed: {payload}")
    return payload

def _docling_markdown(path: Path) -> str:
    module = sys.modules.get("docling.document_converter")
    if module is not None and getattr(module, "__spec__", None) is None:
        from docling.document_converter import DocumentConverter
        return _run_with_timeout(DocumentConverter().convert, str(path)).document.export_to_markdown()
    return _run_conversion_process(_docling_process_entry, path)


def _text_hash(text: str, length: int = 8) -> str:
    """Generate a short stable hash from text for use as a paragraph anchor."""
    return hashlib.sha256(text.encode("utf-8")).hexdigest()[:length]


def _extract_anchors_from_markdown(markdown: str) -> list[dict[str, Any]]:
    """Extract paragraph anchors from Markdown text.

    Uses headings and first sentence of each paragraph as anchors.
    """
    anchors: list[dict[str, Any]] = []
    lines = markdown.split("\n")
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        # Headings
        if stripped.startswith("#"):
            text = stripped.lstrip("#").strip()
            if text:
                anchors.append({"text": text[:50], "anchor": _text_hash(text)})
        # List items
        elif stripped.startswith("- ") or stripped.startswith("* "):
            text = stripped[2:].strip()
            if text and len(text) > 10:
                anchors.append({"text": text[:50], "anchor": _text_hash(text)})
        # Paragraphs (non-empty, non-heading, non-list)
        elif len(stripped) > 20:
            anchors.append({"text": stripped[:50], "anchor": _text_hash(stripped)})
    return anchors


def convert_docx_to_markdown(path: Path) -> tuple[str, list[dict[str, Any]]]:
    """Convert a DOCX file to Markdown using Docling Slim.

    Returns:
        (markdown_string, anchors_list)
        anchors_list is a list of dicts with keys "text" and "anchor"
        for each detected paragraph, used for citation jumping.
    """
    try:
        from docling.document_converter import DocumentConverter
    except ImportError:
        raise ImportError(
            "docling is not installed. Run: pip install docling"
        )

    _ensure_disk_space(path)
    logger.info("converting DOCX: %s", path.name)
    start = time.time()

    markdown = _docling_markdown(path)

    # Extract paragraph anchors from the generated markdown
    anchors = _extract_anchors_from_markdown(markdown)

    elapsed = time.time() - start
    logger.info(
        "DOCX conversion done: %s (%d chars, %d anchors, %.1fs)",
        path.name, len(markdown), len(anchors), elapsed,
    )

    return markdown, anchors


# ── XLSX converter ──────────────────────────────────────────────────────────

_MAX_XLSX_ROWS_PER_CHUNK = 50
_MAX_XLSX_COLS_PER_CHUNK = 20


def _format_cell_pair(
    formula_cell: Any,
    value_cell: Any,
) -> str:
    """Format a cell pair (formula + optional cached value) for Markdown.

    Rules:
      - Regular value (no formula) → formatted as before
      - Formula with cached value → formatted value + formula annotation
      - Formula without cached value → formula expression + no-cache marker
      - Genuinely empty cell → empty string
    """
    from openpyxl.cell.cell import TYPE_FORMULA

    is_formula = formula_cell.data_type == TYPE_FORMULA
    formula_expr = str(formula_cell.value) if is_formula else None

    # Check if the value cell holds a cached result.
    # Use a sentinel to distinguish 0 / False from None.
    _SENTINEL = object()
    cached = value_cell.value if value_cell is not None else _SENTINEL
    has_cached = cached is not _SENTINEL and cached is not None

    if not is_formula:
        # Regular (non-formula) cell — format as before.
        return _format_plain_value(formula_cell)

    # ── Formula cell ─────────────────────────────────────────────────────
    number_format = formula_cell.number_format or ""

    if has_cached:
        # Format the cached value with the formula cell's number format.
        # Create a temporary cell-like object so _format_plain_value can
        # read .value and .number_format.
        class _TempCell:
            pass
        tmp = _TempCell()
        tmp.value = cached
        tmp.number_format = number_format
        formatted = _format_plain_value(tmp)
        return f"{formatted}（公式：{formula_expr}）"
    else:
        return f"公式：{formula_expr}（未缓存计算结果）"


def _format_plain_value(cell: Any) -> str:
    """Format a cell's value for Markdown table output.

    Handles dates, percentages, currency, and other formatted values.
    Falls back to the raw value if formatting fails.
    """
    if cell.value is None:
        return ""
    try:
        nf = (cell.number_format or "").lower()
        # Number format contains date patterns
        if nf and any(kw in nf for kw in ("yyyy", "mm", "dd", "日期")):
            try:
                return cell.value.strftime("%Y-%m-%d") if hasattr(cell.value, "strftime") else str(cell.value)
            except (ValueError, AttributeError):
                pass
        # Percentage
        if nf and "%" in nf:
            try:
                return f"{float(cell.value) * 100:.1f}%"
            except (ValueError, TypeError):
                pass
        # Currency / accounting
        if nf and any(kw in nf for kw in ("¥", "$", "￥", "元")):
            try:
                return f"{float(cell.value):.2f}"
            except (ValueError, TypeError):
                pass
    except Exception:
        pass
    return str(cell.value)


def _detect_data_region_from_formula(sheet: Any) -> tuple[int, int, int, int]:
    """Detect the rectangular data region using the formula sheet.

    Unlike the value sheet, the formula sheet retains formula expressions,
    so cells with uncached formulas are not mistaken for empty cells.

    Returns (min_row, min_col, max_row, max_col).
    """
    if sheet.max_row is None or sheet.max_column is None:
        return (1, 1, 0, 0)

    min_row, min_col = sheet.max_row, sheet.max_column
    max_row, max_col = 1, 1
    found = False

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            if cell.value is not None:
                min_row = min(min_row, cell.row)
                min_col = min(min_col, cell.column)
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
                found = True

    if not found:
        return (1, 1, 0, 0)
    return (min_row, min_col, max_row, max_col)


def _sheet_to_markdown_tables(
    formula_ws: Any,
    value_ws: Any,
    sheet_name: str,
) -> list[dict[str, Any]]:
    """Convert a single sheet to one or more Markdown table chunks.

    Uses the formula worksheet for cell values/formulas and the value
    worksheet for cached calculation results.

    Returns a list of dicts with keys: "markdown", "sheet_name", "cell_range".
    Large tables are split into multiple chunks.
    """
    if formula_ws.max_row is None or formula_ws.max_column is None:
        return []

    min_row, min_col, max_row, max_col = _detect_data_region_from_formula(formula_ws)
    if max_row < min_row or max_col < min_col:
        return []

    # Build a list of (row_idx, [cell_values]) for all data rows
    all_rows: list[list[str]] = []
    for row_idx in range(min_row, max_row + 1):
        row_data: list[str] = []
        for col_idx in range(min_col, max_col + 1):
            f_cell = formula_ws.cell(row=row_idx, column=col_idx)
            v_cell = value_ws.cell(row=row_idx, column=col_idx) if value_ws is not None else None
            row_data.append(_format_cell_pair(f_cell, v_cell))
        all_rows.append(row_data)

    if not all_rows:
        return []

    chunks: list[dict[str, Any]] = []
    n_cols = max_col - min_col + 1

    # Split rows into chunks
    for chunk_start in range(0, len(all_rows), _MAX_XLSX_ROWS_PER_CHUNK):
        chunk_end = min(chunk_start + _MAX_XLSX_ROWS_PER_CHUNK, len(all_rows))
        chunk_rows = all_rows[chunk_start:chunk_end]

        # If the chunk table is wider than the limit, split columns too
        if n_cols > _MAX_XLSX_COLS_PER_CHUNK:
            for col_start in range(0, n_cols, _MAX_XLSX_COLS_PER_CHUNK):
                col_end = min(col_start + _MAX_XLSX_COLS_PER_CHUNK, n_cols)
                chunk = _build_table_markdown(chunk_rows, col_start, col_end)
                if chunk.strip():
                    actual_col_start = min_col + col_start
                    actual_col_end = min_col + col_end - 1
                    chunks.append({
                        "markdown": f"## Sheet: {sheet_name}\n\n{chunk}\n",
                        "sheet_name": sheet_name,
                        "cell_range": f"{_col_letter(actual_col_start)}{min_row + chunk_start}:{_col_letter(actual_col_end)}{min_row + chunk_end - 1}",
                    })
        else:
            chunk = _build_table_markdown(chunk_rows, 0, n_cols)
            if chunk.strip():
                actual_col_start = min_col
                actual_col_end = min_col + n_cols - 1
                chunks.append({
                    "markdown": f"## Sheet: {sheet_name}\n\n{chunk}\n",
                    "sheet_name": sheet_name,
                    "cell_range": f"{_col_letter(actual_col_start)}{min_row + chunk_start}:{_col_letter(actual_col_end)}{min_row + chunk_end - 1}",
                })

    return chunks


def _build_table_markdown(rows: list[list[str]], col_start: int, col_end: int) -> str:
    """Build a Markdown table from a slice of rows."""
    if not rows:
        return ""

    # Header row
    header = rows[0][col_start:col_end]
    result = "| " + " | ".join(header) + " |\n"
    result += "| " + " | ".join("---" for _ in header) + " |\n"

    # Data rows
    for row in rows[1:]:
        cells = row[col_start:col_end]
        result += "| " + " | ".join(cells) + " |\n"

    return result


def _col_letter(n: int) -> str:
    """Convert a 1-indexed column number to Excel column letters (A, B, ..., Z, AA, ...)."""
    letters = ""
    while n > 0:
        n -= 1
        letters = chr(65 + n % 26) + letters
        n //= 26
    return letters


def recalculate_xlsx(path: Path) -> Path:
    """Recalculate XLSX formulas via LibreOffice service.

    Returns the path to a temporary file with cached formula values.
    The caller is responsible for cleanup.
    """
    import httpx
    from .config import LIBREOFFICE_URL, LIBREOFFICE_TIMEOUT

    logger.info("recalculating XLSX via LibreOffice: %s", path.name)
    start = time.time()

    with httpx.Client(timeout=LIBREOFFICE_TIMEOUT) as client:
        with open(path, "rb") as fh:
            resp = client.post(
                f"{LIBREOFFICE_URL}/v1/recalculate",
                files={"file": (path.name, fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        if resp.status_code != 200:
            raise RuntimeError(
                f"LibreOffice recalculation failed (HTTP {resp.status_code}): {resp.text[:200]}"
            )

    # Save the recalculated file to a temp location
    temp = path.parent / (path.stem + ".recalculated" + path.suffix)
    temp.write_bytes(resp.content)

    elapsed = time.time() - start
    logger.info("recalculation done: %s (%.1fs)", path.name, elapsed)
    return temp


def _convert_xlsx_to_markdown_impl(path: Path) -> tuple[str, list[dict[str, Any]]]:
    """Convert an XLSX file to Markdown using openpyxl.

    Loads the workbook twice — once for formulas (data_only=False) and once
    for cached values (data_only=True) — so that formula cells without a
    cached result are not silently dropped.

    Returns:
        (markdown_string, sheets_metadata)
        sheets_metadata is a list of dicts with keys "sheet_name" and "cell_range"
        for each chunk, used for citation jumping.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is not installed. Run: pip install openpyxl"
        )
    _ensure_disk_space(path)

    logger.info("converting XLSX: %s", path.name)
    start = time.time()

    formula_wb = openpyxl.load_workbook(path, data_only=False, read_only=False)
    value_wb = openpyxl.load_workbook(path, data_only=True, read_only=False)

    uncached_count = 0
    chunks: list[dict[str, Any]] = []

    try:
        for sheet_name in formula_wb.sheetnames:
            formula_ws = formula_wb[sheet_name]
            # Skip hidden sheets
            if formula_ws.sheet_state in ("hidden", "veryHidden"):
                logger.info("  skipping hidden sheet: %s", sheet_name)
                continue
            value_ws = value_wb[sheet_name] if sheet_name in value_wb.sheetnames else None
            sheet_chunks = _sheet_to_markdown_tables(formula_ws, value_ws, sheet_name)
            chunks.extend(sheet_chunks)

            # Count uncached formula cells for aggregate logging
            for row in formula_ws.iter_rows(min_row=1, max_row=formula_ws.max_row or 0, max_col=formula_ws.max_column or 0):
                for cell in row:
                    from openpyxl.cell.cell import TYPE_FORMULA
                    if cell.data_type == TYPE_FORMULA and cell.value is not None:
                        v = value_ws.cell(row=cell.row, column=cell.column).value if value_ws else None
                        if v is None:
                            uncached_count += 1
    finally:
        formula_wb.close()
        value_wb.close()

    if uncached_count:
        logger.info(
            "XLSX conversion: %d formula cells have no cached result",
            uncached_count,
        )

    # Combine all chunks into a single markdown string
    markdown = "\n".join(c["markdown"] for c in chunks)

    # Extract metadata for citation jumping
    sheets_metadata = [
        {"sheet_name": c["sheet_name"], "cell_range": c["cell_range"]}
        for c in chunks
    ]

    elapsed = time.time() - start
    logger.info(
        "XLSX conversion done: %s (%d sheets, %d chunks, %d chars, %.1fs)",
        path.name, len(formula_wb.sheetnames), len(chunks), len(markdown), elapsed,
    )

    return markdown, sheets_metadata


def convert_xlsx_to_markdown(path: Path) -> tuple[str, list[dict[str, Any]]]:
    """Convert XLSX in a child process so timeout termination releases resources."""
    return _run_conversion_process(_xlsx_process_entry, path)


# ── PPTX converter ──────────────────────────────────────────────────────────


def convert_pptx_to_markdown(path: Path) -> tuple[str, list[dict[str, Any]]]:
    """Convert a PPTX file to Markdown using Docling Slim.

    Each slide is separated by a ``---`` marker in the output.
    The anchors list includes slide numbers for citation jumping.

    Returns:
        (markdown_string, slides_metadata)
        slides_metadata is a list of dicts with keys "slide_number" and "text".
    """
    try:
        from docling.document_converter import DocumentConverter
    except ImportError:
        raise ImportError(
            "docling is not installed. Run: pip install docling"
        )

    logger.info("converting PPTX: %s", path.name)
    start = time.time()

    _ensure_disk_space(path)
    markdown = _docling_markdown(path)

    # Count slides by heading markers
    slides: list[dict[str, Any]] = []
    slide_num = 0
    for line in markdown.split("\n"):
        stripped = line.strip()
        if stripped.startswith("#") and len(stripped) > 10:
            slide_num += 1
            slides.append({
                "slide_number": slide_num,
                "text": stripped.lstrip("#").strip()[:50],
            })

    slide_count = max(slide_num, 1)

    elapsed = time.time() - start
    logger.info(
        "PPTX conversion done: %s (%d slides, %d chars, %.1fs)",
        path.name, slide_count, len(markdown), elapsed,
    )

    return markdown, slides


def is_valid_pdf_file(path: Path) -> bool:
    """Return whether a regular file has a PDF signature."""
    try:
        if not path.is_file() or path.is_symlink():
            return False
        with path.open("rb") as handle:
            return handle.read(5) == b"%PDF-"
    except OSError:
        return False


def convert_pptx_to_pdf(path: Path) -> Path:
    """Convert a PPTX file to PDF using LibreOffice service.

    Returns the path to the generated PDF file.
    The caller is responsible for cleanup.
    """
    import httpx
    from .config import LIBREOFFICE_URL, LIBREOFFICE_TIMEOUT

    logger.info("converting PPTX to PDF via LibreOffice: %s", path.name)
    start = time.time()

    with httpx.Client(timeout=LIBREOFFICE_TIMEOUT) as client:
        with open(path, "rb") as fh:
            resp = client.post(
                f"{LIBREOFFICE_URL}/v1/convert",
                params={"target_format": "pdf"},
                files={"file": (path.name, fh, "application/vnd.openxmlformats-officedocument.presentationml.presentation")},
            )
        if resp.status_code != 200:
            if resp.status_code == 413:
                raise PptxPreviewFileTooLargeError(
                    f"PPTX to PDF conversion rejected oversized input (HTTP 413): {resp.text[:200]}"
                )
            raise RuntimeError(
                f"PPTX to PDF conversion failed (HTTP {resp.status_code}): {resp.text[:200]}"
            )
        if not resp.content.startswith(b"%PDF-"):
            raise RuntimeError("PPTX to PDF conversion failed: invalid PDF output")

    # Write beside the final artifact so os.replace remains atomic on the
    # mounted production filesystem. A failed write never damages an existing
    # valid preview.
    pdf_path = path.with_suffix(".preview.pdf")
    temporary_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="wb",
            prefix=f".{pdf_path.name}.",
            suffix=".tmp",
            dir=pdf_path.parent,
            delete=False,
        ) as handle:
            temporary_path = Path(handle.name)
            handle.write(resp.content)
            handle.flush()
            os.fsync(handle.fileno())
        if not is_valid_pdf_file(temporary_path):
            raise RuntimeError("PPTX to PDF conversion failed: invalid PDF output")
        os.replace(temporary_path, pdf_path)
    finally:
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)

    elapsed = time.time() - start
    logger.info("PPTX to PDF done: %s (%.1fs)", path.name, elapsed)
    return pdf_path


def _md_path_for_office(source_path: Path, parsed_dir: Path) -> Path:
    """Generate the cached Markdown path for an Office document.

    Uses the same naming convention as MinerU: relative path under DOCS_DIR
    with path separators replaced by double underscores.
    """
    from src.config import DOCS_DIR
    try:
        rel = source_path.relative_to(DOCS_DIR)
    except ValueError:
        rel = Path(source_path.name)
    stem = rel.with_suffix("").as_posix().replace("/", "__")
    return parsed_dir / f"{stem}.md"
