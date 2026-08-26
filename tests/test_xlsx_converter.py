"""Tests for XLSX → Markdown converter, focusing on formula cell handling.

Run with:
    pytest tests/test_xlsx_converter.py -v
"""

from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace

import pytest

from src.office_convert import (
    _format_cell_pair,
    _format_plain_value,
    _detect_data_region_from_formula,
    convert_xlsx_to_markdown,
)
from src.document_locations import read_location_sidecar
from src.indexing_pipeline import _build_xlsx_doc


# ── _format_plain_value ─────────────────────────────────────────────────────

class TestFormatPlainValue:
    """Test _format_plain_value with various cell types."""

    def test_plain_string(self):
        cell = _cell(value="hello")
        assert _format_plain_value(cell) == "hello"

    def test_plain_integer(self):
        cell = _cell(value=42)
        assert _format_plain_value(cell) == "42"

    def test_plain_float(self):
        cell = _cell(value=3.14)
        assert _format_plain_value(cell) == "3.14"

    def test_none(self):
        cell = _cell(value=None)
        assert _format_plain_value(cell) == ""

    def test_percentage(self):
        cell = _cell(value=0.13, number_format="0.0%")
        assert _format_plain_value(cell) == "13.0%"

    def test_currency(self):
        cell = _cell(value=420.50, number_format="¥#,##0.00")
        assert _format_plain_value(cell) == "420.50"

    def test_zero(self):
        cell = _cell(value=0)
        assert _format_plain_value(cell) == "0"


# ── _format_cell_pair ───────────────────────────────────────────────────────

class TestFormatCellPair:
    """Test _format_cell_pair with formula and value cell pairs."""

    def test_regular_value(self):
        f = _cell(value="hello", data_type="s")
        v = _cell(value="hello")
        assert _format_cell_pair(f, v) == "hello"

    def test_formula_with_cached_value(self):
        f = _cell(value="=E4*F4", data_type="f")
        v = _cell(value=52814.80)
        result = _format_cell_pair(f, v)
        assert "52814.8" in result
        assert "=E4*F4" in result

    def test_formula_without_cached_value(self):
        f = _cell(value="=E4*F4", data_type="f")
        v = _cell(value=None)
        result = _format_cell_pair(f, v)
        assert "=E4*F4" in result
        assert "未缓存" in result

    def test_formula_cached_zero(self):
        f = _cell(value="=1-1", data_type="f")
        v = _cell(value=0)
        result = _format_cell_pair(f, v)
        assert "0" in result
        assert "=1-1" in result

    def test_formula_cached_false(self):
        f = _cell(value="=1=0", data_type="f")
        v = _cell(value=False)
        result = _format_cell_pair(f, v)
        # False is a valid cached value; should not be treated as empty
        assert "False" in result or "false" in result

    def test_empty_cell(self):
        f = _cell(value=None)
        v = _cell(value=None)
        assert _format_cell_pair(f, v) == ""

    def test_formula_with_percentage_cached(self):
        f = _cell(value="=H4*(1+G4)", data_type="f", number_format="0.0%")
        v = _cell(value=0.15)
        result = _format_cell_pair(f, v)
        assert "15.0%" in result or "0.15" in result
        assert "=H4*(1+G4)" in result


# ── _detect_data_region_from_formula ────────────────────────────────────────

class TestDetectDataRegionFromFormula:
    """Test data region detection using the formula sheet."""

    def test_region_with_formulas(self, xlsx_with_formulas):
        """Cells with uncached formulas should still be included in the region."""
        wb = _open_xlsx(xlsx_with_formulas, data_only=False)
        ws = wb["材料参数"]
        min_r, min_c, max_r, max_c = _detect_data_region_from_formula(ws)
        wb.close()
        # Should include the formula columns (H and I)
        assert max_c >= 9  # Column I = 9
        assert max_r >= 4  # Row 4 has data

    def test_region_empty_sheet(self, xlsx_empty):
        wb = _open_xlsx(xlsx_empty, data_only=False)
        ws = wb.active
        min_r, min_c, max_r, max_c = _detect_data_region_from_formula(ws)
        wb.close()
        assert max_r == 0 or max_c == 0


# ── convert_xlsx_to_markdown ────────────────────────────────────────────────

class TestConvertXlsxToMarkdown:
    """Integration tests for the full XLSX converter."""

    def test_formulas_in_output(self, xlsx_with_formulas):
        """Formula cells should appear in the Markdown output."""
        md, meta = convert_xlsx_to_markdown(xlsx_with_formulas)
        assert "=E4*F4" in md
        assert "未缓存" in md

    def test_plain_values_preserved(self, xlsx_with_formulas):
        """Regular (non-formula) values should still appear."""
        md, meta = convert_xlsx_to_markdown(xlsx_with_formulas)
        assert "MAT-001" in md
        assert "420.50" in md or "420.5" in md
        assert "125.6" in md

    def test_percentage_preserved(self, xlsx_with_formulas):
        """Percentage values should be formatted correctly."""
        md, meta = convert_xlsx_to_markdown(xlsx_with_formulas)
        assert "13.0%" in md or "13%" in md

    def test_hidden_sheet_skipped(self, xlsx_with_hidden_sheet):
        """Hidden sheets should not appear in the output."""
        md, meta = convert_xlsx_to_markdown(xlsx_with_hidden_sheet)
        assert "隐藏" not in md

    def test_chunk_metadata(self, xlsx_with_formulas):
        """Sheet chunks should preserve sheet_name and cell_range."""
        md, meta = convert_xlsx_to_markdown(xlsx_with_formulas)
        assert any(m["sheet_name"] == "材料参数" for m in meta)
        for m in meta:
            assert ":" in m["cell_range"]  # Should be like A1:H4

    def test_index_sidecar_preserves_sheet_section_anchor(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ):
        source = tmp_path / "sheet.xlsx"
        source.write_bytes(b"synthetic")
        markdown = "## Sheet: 统计\n\n| 编号 |\n| --- |\n| A |\n"
        monkeypatch.setattr(
            "src.indexing_pipeline.recalculate_xlsx",
            lambda _path: (_ for _ in ()).throw(RuntimeError("unavailable")),
        )
        monkeypatch.setattr(
            "src.indexing_pipeline.convert_xlsx_to_markdown",
            lambda _path: (
                markdown,
                [{
                    "text": markdown,
                    "sheet_name": "统计",
                    "cell_range": "A1:A2",
                }],
            ),
        )

        document = _build_xlsx_doc(
            source,
            lambda _status: None,
            parsed_dir=tmp_path / "parsed",
            write_preview=False,
            force_parse=True,
        )
        locations = read_location_sidecar(document.location_map_path)

        assert locations[0].heading_anchor == "Sheet: 统计"


# ── Fixtures ─────────────────────────────────────────────────────────────────

def _cell(value=None, data_type="s", number_format=None):
    """Create a minimal mock cell for testing."""
    return SimpleNamespace(
        value=value,
        data_type=data_type,
        number_format=number_format,
    )


def _open_xlsx(path, data_only=False):
    import openpyxl
    return openpyxl.load_workbook(path, data_only=data_only)


@pytest.fixture
def xlsx_with_formulas(tmp_path: Path):
    """Create a minimal XLSX with formulas (no cached values)."""
    import openpyxl
    path = tmp_path / "test_formulas.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "材料参数"
    ws.append(["编号", "材料名称", "规格", "备注", "单价", "数量", "税率", "未税金额", "含税金额"])
    ws.append([])
    ws.append([])
    ws.append(["MAT-001", "普通混凝土", "C30", "合成样本", 420.50, 125.6, 0.13, None, None])
    ws["G4"].number_format = "0.0%"
    ws["H4"] = 420.50 * 125.6
    ws["I4"] = 52814.80 * (1 + 0.13)
    wb.save(path)
    wb.close()

    # Reopen and replace H4 with an uncached formula while retaining I4 as a value.
    wb2 = openpyxl.load_workbook(path)
    ws2 = wb2.active
    ws2["H4"] = "=E4*F4"
    wb2.save(path)
    wb2.close()
    return path


@pytest.fixture
def xlsx_empty(tmp_path: Path):
    """Create an empty XLSX."""
    import openpyxl
    path = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def xlsx_with_hidden_sheet(tmp_path: Path):
    """Create an XLSX with a hidden sheet."""
    import openpyxl
    path = tmp_path / "hidden.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "可见"
    ws.append(["data"])
    ws2 = wb.create_sheet("隐藏")
    ws2.sheet_state = "hidden"
    ws2.append(["should not appear"])
    wb.save(path)
    wb.close()
    return path
